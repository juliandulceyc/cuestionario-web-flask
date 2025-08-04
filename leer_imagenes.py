# CREAR leer_imagenes_reales.py
import pandas as pd
from openpyxl import load_workbook
import base64
import os

def extraer_imagenes_reales():
    """Extrae las imágenes REALES del Excel, ignorando el texto 'Imagen'"""
    
    archivo = 'Evaluación FWS PAN V2.xlsx'
    
    print("🔍 EXTRAYENDO IMÁGENES REALES (ignorando texto 'Imagen')")
    print("=" * 60)
    
    if not os.path.exists(archivo):
        print(f"❌ Archivo {archivo} no encontrado")
        return
    
    try:
        # 1. Cargar con openpyxl para acceder a imágenes
        wb = load_workbook(archivo)
        ws = wb.active
        
        print(f"📊 Archivo: {archivo}")
        print(f"   Hojas: {wb.sheetnames}")
        print(f"   Hoja activa: {ws.title}")
        print(f"   Dimensiones: {ws.max_row} x {ws.max_column}")
        
        # 2. Mapear imágenes por posición
        imagenes_por_posicion = {}
        
        if hasattr(ws, '_images') and ws._images:
            print(f"\n🖼️ IMÁGENES ENCONTRADAS: {len(ws._images)}")
            
            for i, img in enumerate(ws._images):
                try:
                    # Obtener posición de la imagen
                    if hasattr(img.anchor, '_from'):
                        fila = img.anchor._from.row + 1  # Convertir a base 1
                        columna = img.anchor._from.col + 1
                        
                        print(f"\n   📸 Imagen {i+1}:")
                        print(f"      Posición: Fila {fila}, Columna {columna}")
                        print(f"      Tipo: {type(img)}")
                        
                        # Extraer datos de la imagen
                        img_data = None
                        if hasattr(img, '_data'):
                            img_data = img._data()
                        elif hasattr(img, 'data'):
                            img_data = img.data
                        
                        if img_data and len(img_data) > 100:
                            # Convertir a base64
                            imagen_base64 = base64.b64encode(img_data).decode()
                            
                            # Detectar tipo de imagen
                            if img_data.startswith(b'\x89PNG'):
                                mime_type = 'png'
                            elif img_data.startswith(b'\xFF\xD8\xFF'):
                                mime_type = 'jpeg'
                            elif img_data.startswith(b'GIF8'):
                                mime_type = 'gif'
                            else:
                                mime_type = 'png'
                            
                            data_url = f"data:image/{mime_type};base64,{imagen_base64}"
                            
                            # Guardar por posición
                            key = f"{fila}-{columna}"
                            imagenes_por_posicion[key] = {
                                'fila': fila,
                                'columna': columna,
                                'data_url': data_url,
                                'tamaño': len(img_data),
                                'tipo': mime_type
                            }
                            
                            print(f"      ✅ Extraída: {mime_type} ({len(img_data):,} bytes)")
                            print(f"      Base64: {len(imagen_base64)} caracteres")
                        else:
                            print(f"      ❌ Sin datos válidos")
                            
                except Exception as e:
                    print(f"      ❌ Error procesando imagen {i+1}: {e}")
        else:
            print(f"\n❌ NO SE ENCONTRARON IMÁGENES")
            return {}
        
        print(f"\n📊 RESUMEN:")
        print(f"   Total imágenes extraídas: {len(imagenes_por_posicion)}")
        
        return imagenes_por_posicion
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return {}

def crear_excel_con_imagenes_reales():
    """Crear nuevo Excel con las imágenes como data URLs en las celdas"""
    
    archivo_original = 'Evaluación FWS PAN V2.xlsx'
    archivo_nuevo = 'Evaluación_con_imagenes_reales.xlsx'
    
    print(f"\n🔧 CREANDO EXCEL CON IMÁGENES REALES")
    print("=" * 50)
    
    # 1. Extraer imágenes reales
    imagenes = extraer_imagenes_reales()
    
    if not imagenes:
        print("❌ No se pudieron extraer imágenes")
        return None
    
    # 2. Cargar datos con pandas
    try:
        df = pd.read_excel(archivo_original)
        print(f"\n📊 Excel cargado: {len(df)} filas")
        
        # 3. Determinar qué columna es IMAGEN
        columna_imagen_index = None
        if 'IMAGEN' in df.columns:
            columna_imagen_index = df.columns.get_loc('IMAGEN') + 1  # +1 para base 1
            print(f"   Columna IMAGEN: índice {columna_imagen_index}")
        else:
            print("❌ Columna IMAGEN no encontrada")
            return None
        
        # 4. Reemplazar texto "Imagen" con data URLs reales
        imagenes_asignadas = 0
        
        for index, row in df.iterrows():
            fila_excel = index + 2  # pandas index 0 = fila 2 en Excel
            key = f"{fila_excel}-{columna_imagen_index}"
            
            # Si hay imagen en esta posición
            if key in imagenes:
                imagen_info = imagenes[key]
                
                # Reemplazar el valor de la celda con la data URL
                df.at[index, 'IMAGEN'] = imagen_info['data_url']
                
                pregunta = str(row.get('PREGUNTA', ''))[:50]
                print(f"   📝 Fila {fila_excel}: {pregunta}... → {imagen_info['tipo']} ({imagen_info['tamaño']:,} bytes)")
                
                imagenes_asignadas += 1
        
        # 5. Guardar nuevo Excel
        df.to_excel(archivo_nuevo, index=False)
        
        print(f"\n✅ ÉXITO:")
        print(f"   Archivo creado: {archivo_nuevo}")
        print(f"   Imágenes convertidas: {imagenes_asignadas}")
        
        return archivo_nuevo
        
    except Exception as e:
        print(f"❌ Error creando Excel: {e}")
        import traceback
        traceback.print_exc()
        return None

def verificar_excel_final(archivo):
    """Verificar que el Excel final tiene las imágenes como data URLs"""
    
    print(f"\n🔍 VERIFICANDO EXCEL FINAL: {archivo}")
    print("=" * 50)
    
    try:
        df = pd.read_excel(archivo)
        imagenes = df['IMAGEN'].dropna()
        
        imagenes_validas = 0
        
        for i, valor in enumerate(imagenes.head(10)):
            valor_str = str(valor)
            
            print(f"\n   Imagen {i+1}:")
            print(f"      Tipo: {type(valor)}")
            print(f"      Longitud: {len(valor_str):,} caracteres")
            
            if valor_str.startswith('data:image'):
                imagenes_validas += 1
                print(f"      ✅ IMAGEN VÁLIDA: {valor_str[:50]}...")
                
                # Verificar que el base64 es decodificable
                try:
                    base64_part = valor_str.split(';base64,')[1]
                    base64.b64decode(base64_part[:100])  # Solo probar primeros 100 chars
                    print(f"      ✅ Base64 válido")
                except:
                    print(f"      ❌ Base64 inválido")
            else:
                print(f"      ❌ No es imagen: {valor_str[:50]}...")
        
        print(f"\n📊 RESULTADO FINAL:")
        print(f"   Total valores en IMAGEN: {len(imagenes)}")
        print(f"   Imágenes válidas: {imagenes_validas}")
        
        if imagenes_validas > 0:
            print(f"   ✅ ÉXITO: Excel listo para usar en app.py")
            return True
        else:
            print(f"   ❌ No hay imágenes válidas")
            return False
            
    except Exception as e:
        print(f"❌ Error verificando: {e}")
        return False

if __name__ == "__main__":
    print("🚀 SOLUCIÓN: EXTRAER IMÁGENES REALES DE EXCEL")
    print("=" * 60)
    
    # Proceso completo
    archivo_final = crear_excel_con_imagenes_reales()
    
    if archivo_final:
        if verificar_excel_final(archivo_final):
            print(f"\n🎯 PROCESO COMPLETADO:")
            print(f"   ✅ Archivo listo: {archivo_final}")
            print(f"   📋 Próximo paso: En app.py línea 162, cambiar:")
            print(f"   archivo_excel = '{archivo_final}'")
        else:
            print(f"\n❌ El archivo final no es válido")
    else:
        print(f"\n❌ No se pudo crear el archivo final")